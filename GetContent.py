from bs4 import BeautifulSoup
import urllib.request
import fake_useragent
import json
import requests
import pathlib
import os
import string
import random
import openpyxl
import datetime




class GetContent():
    ua = fake_useragent.UserAgent()
    def __init__(self):
        self.headers = {"User-Agent": self.ua.random}
        self.propertiesName = []
        self.products = {}

    def randomStr(self, nrandchars):
        alpha = string.ascii_letters + string.digits
        chars = ''.join(random.choice(alpha) for _ in range(nrandchars))
        return chars

    def get_files(self, link, folder='images'):
        response = requests.get(link, stream=True)
        ext = pathlib.Path(link).suffix
        if not os.path.isdir("upload"):
            os.mkdir("upload")
        if not os.path.isdir("upload/" + folder):
            os.mkdir("upload/" + folder)
        file_name = 'upload/' + folder + '/' + folder + '_' + self.randomStr(10) + ext
        file = open(file_name, 'bw')
        for chunk in response.iter_content(4096):
            file.write(chunk)
        return file_name

    # сылки пагинации
    def getBreadcrumbs(self, url):
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            exit()
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        breadcrumbsContent = soup.find_all('a', {'class': 'ty-pagination__item'})
        breadcrumbsLinks = []
        breadcrumbsLinks.append(url)
        if len(breadcrumbsContent) > 0:
            for breadcrumb in breadcrumbsContent:
                try:
                    pageLink = breadcrumb['href']
                    if pageLink not in breadcrumbsLinks:
                        breadcrumbsLinks.append(pageLink)
                except:
                    pass
        return breadcrumbsLinks

    #сылки с одной страницы списка товаров
    def getLinksOnePage(self, url):
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            exit()
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        pegeList = soup.find_all('a', {'class': 'product-title'})
        pegeListLinks = []
        if len(pegeList) > 0:
            for pageLink in pegeList:
                pegeListLinks.append(pageLink['href'])
        return pegeListLinks

    # контент карточки товара и файли картинок
    def getItem(self, url):
        item = dict()
        item['url'] = url
        item['title'] = None
        item['price'] = 0
        item['main'] = None
        item['more'] = ''
        item['description'] = ''
        item['video'] = ''
        item['json'] = None
        item['attributes'] = None
        propertiesName = []
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            exit()
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        itemContent = soup.find('div', {'class': 'ty-product-bigpicture'})
        item['title'] = itemContent.find('h1', {'class':'ty-product-block-title'}).get_text(strip=True)
        item['price'] = itemContent.find('div', {'class':'ty-product-block__price-actual'}).find('span', {'class': 'ty-price-num'}).get_text(strip=True)
        try:
            text = itemContent.find('div', {'id':'content_description'}).decode_contents().replace('\n', '').replace('\r', '').replace('<div>', '').replace('</div>', '')
            text = text.replace('<p style="color:red;">*Внешний вид товара может незначительно отличаться от фотографий на сайте</p>', '')
            item['description'] = text
        except:
            pass
        try:
            properties = itemContent.find('div', {'id':'content_features'}).find_all('div', {'class':'ty-product-feature'})
            if len(properties) > 0:
                atrr = []
                for props in properties:
                    element = {
                        'label': props.find('div', class_='ty-product-feature__label').get_text(strip=True),
                        'value': props.find('div', class_='ty-product-feature__value').get_text(strip=True)
                    }
                    if element['label'] not in self.propertiesName:
                        self.propertiesName.append(element['label'])
                    atrr.append(element)
                if len(atrr) > 0:
                    forJson = {}
                    key = 1
                    for el in atrr:
                        forJson[key] = el
                        key += 1
                    item['json'] = json.dumps(forJson, ensure_ascii=False)

                item['attributes'] = atrr
        except:
            pass
        try:
            mainImgLink = soup.find('div', class_='ty-product-img').a['href']
            item['main'] = self.get_files(mainImgLink)
        except:
            pass
        try:
            moreImgLinksBlock = soup.find_all('a', class_='cm-image-previewer')
            if len(moreImgLinksBlock) > 1:
                moreLinks = []
                for moreLink in moreImgLinksBlock[1:]:
                    try:
                        path = self.get_files(moreLink['href'], 'more')
                        moreLinks.append(path)
                    except:
                        pass
                if len(moreLinks) > 0:
                    item['more'] = ','.join(moreLinks)
        except:
            pass
        try:
            videoItems = soup.find_all('div', class_='cp-video-grid__item-name')
            if len(videoItems) > 0:
                videoData = []
                for video in videoItems:
                    videoItem = {
                        'name': video.find('a').get_text(strip=True),
                        'link': video.find('a')['href']
                    }
                    videoData.append(videoItem)
                    item['video'] = json.dumps(videoData, ensure_ascii=False)
        except:
            pass
        return item

    def getExel(self, data):
        book = openpyxl.Workbook()
        sheet = book.active

        sheet.cell(row=1, column=1).value = 'URL'
        sheet.cell(row=1, column=2).value = 'NAME'
        sheet.cell(row=1, column=3).value = 'PRICE'
        sheet.cell(row=1, column=4).value = 'DESCRIPTION'
        sheet.cell(row=1, column=5).value = 'MAIN_IMG'
        sheet.cell(row=1, column=6).value = 'MORE_IMG'
        sheet.cell(row=1, column=7).value = 'VIDEO'
        sheet.cell(row=1, column=8).value = 'JSON_PROPERTIES'



        maxColumn = 9 + len(self.propertiesName)
        if maxColumn > 9:
            propNumber = 9
            for name in self.propertiesName:
                sheet.cell(row=1, column=propNumber).value = name
                propNumber += 1

        rowNumber = 2

        for item in data:
            sheet.cell(row=rowNumber, column=1).value = item['url']
            sheet.cell(row=rowNumber, column=2).value = item['title']
            sheet.cell(row=rowNumber, column=3).value = item['price']
            sheet.cell(row=rowNumber, column=4).value = item['description']
            sheet.cell(row=rowNumber, column=5).value = item['main']
            sheet.cell(row=rowNumber, column=6).value = item['more']
            sheet.cell(row=rowNumber, column=7).value = item['video']
            sheet.cell(row=rowNumber, column=8).value = item['json']
            if item['attributes'] is not None:
                for col in range(9, maxColumn):
                    for atr in item['attributes']:
                        if sheet.cell(row=1, column=col).value == atr['label']:
                            sheet.cell(row=rowNumber, column=col).value = atr['value']
                        else:
                            pass
            rowNumber += 1

        uniq_filename = str(datetime.datetime.now().date()) + '_' + str(datetime.datetime.now().time()).replace(':', '_')
        book.save('resault_' + uniq_filename + '.xlsx')
        book.close()
        print('************ ------ *************')


        # сылки на детальные страницы Видеообзоров с одной страницы списка
    def getLinksFromOnePageOfList(self, url):
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            exit()
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        linkWrappers = soup.find_all('div', {'class': 'cp-video-grid__item-name'})
        pegeListLinks = []
        if linkWrappers is not None:
            for pageLink in linkWrappers:
                itemUrl = pageLink.a['href']
                pegeListLinks.append(itemUrl)
        return pegeListLinks

    # Контент с детальной страницы видеообзора
    def getVideoItem(self, url):
        item = dict()
        item['url'] = url
        item['title'] = None
        item['link'] = None
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            exit()
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        item['title'] = soup.find('h1', {'class': 'ty-mainbox-title'}).get_text(strip=True)
        link = soup.find('div', {'class': 'cp-video-detailed__video-wrapper'}).iframe['src']
        if link.find('?') != -1:
            item['link'] = link.split('?')[0]
        else:
            item['link'] = link
        return item

    def vieoToExel(self, data):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(row=1, column=1).value = 'URL'
        sheet.cell(row=1, column=2).value = 'TITLE'
        sheet.cell(row=1, column=3).value = 'LINK'

        rowNumber = 2
        for item in data:
            sheet.cell(row=rowNumber, column=1).value = item['url']
            sheet.cell(row=rowNumber, column=2).value = item['title']
            sheet.cell(row=rowNumber, column=3).value = item['link']
            rowNumber +=1

        uniq_filename = str(datetime.datetime.now().date()) + '_' + str(datetime.datetime.now().time()).replace(':','_')
        book.save('video_' + uniq_filename + '.xlsx')
        book.close()
        print('************ ------ *************')

